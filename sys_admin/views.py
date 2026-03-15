from django.shortcuts import render, get_object_or_404, redirect
from accounts.decorators import admin_required
from property.models import Property
from booking.models import BookingRequest
from accounts.models import User, KYC, ProMembership
from django.contrib import messages
from django.http import HttpResponse
import openpyxl
from datetime import datetime

@admin_required
def dashboard(request):
    """System Admin Overview Dashboard."""
    # Summary Stats
    total_tenants = User.objects.filter(role='TENANT').count()
    total_owners = User.objects.filter(role='OWNER').count()
    total_properties = Property.objects.count()
    total_bookings = BookingRequest.objects.count()
    
    # Recent Activities (Placeholder for now)
    recent_properties = Property.objects.order_by('-created_at')[:5]
    
    context = {
        'total_tenants': total_tenants,
        'total_owners': total_owners,
        'total_properties': total_properties,
        'total_bookings': total_bookings,
        'recent_properties': recent_properties,
    }
    return render(request, 'sys_admin/dashboard.html', context)
    
@admin_required
def user_matrix(request):
    """Deep dive into user demographics and verification status."""
    # Tenant Stats
    tenants = User.objects.filter(role='TENANT')
    tenant_stats = {
        'total': tenants.count(),
        'active': tenants.filter(is_active=True).count(),
        'verified': tenants.filter(kyc__is_verified=True).count(),
        'pending_kyc': tenants.filter(kyc__is_verified=False).count(),
    }
    
    # Owner Stats
    owners = User.objects.filter(role='OWNER')
    owner_stats = {
        'total': owners.count(),
        'active': owners.filter(is_active=True).count(),
        'verified': owners.filter(kyc__is_verified=True).count(),
        'pending_kyc': owners.filter(kyc__is_verified=False).count(),
    }
    
    return render(request, 'sys_admin/user_matrix.html', {
        'tenant_stats': tenant_stats,
        'owner_stats': owner_stats,
    })

@admin_required
def listing_requests(request):
    """View to manage property verification requests."""
    properties = Property.objects.filter(is_verified=False).order_by('-created_at')
    return render(request, 'sys_admin/operations/listing_requests.html', {'properties': properties})

@admin_required
def tenant_kyc(request):
    """View to manage tenant KYC verification."""
    # Show all pending KYC records to ensure nothing is missed
    kyc_list = KYC.objects.filter(is_verified=False).order_by('-submitted_at')
    # Filter by tenants in the template or here if needed, but let's see all for now
    return render(request, 'sys_admin/operations/tenant_kyc.html', {'kyc_list': kyc_list})

@admin_required
def owner_kyc(request):
    """View to manage owner KYC verification."""
    # Show all pending KYC records for visibility
    kyc_list = KYC.objects.filter(is_verified=False).order_by('-submitted_at')
    return render(request, 'sys_admin/operations/owner_kyc.html', {'kyc_list': kyc_list})

@admin_required
def tenant_kyc_detail(request, pk):
    """Detailed audit view for a specific tenant's KYC."""
    kyc = get_object_or_404(KYC, pk=pk, user__role='TENANT')
    return render(request, 'sys_admin/operations/tenant_kyc_detail.html', {'kyc': kyc})

@admin_required
def owner_kyc_detail(request, pk):
    """Detailed audit view for a specific owner's KYC."""
    kyc = get_object_or_404(KYC, pk=pk, user__role='OWNER')
    return render(request, 'sys_admin/operations/owner_kyc_detail.html', {'kyc': kyc})

@admin_required
def approve_kyc(request, pk):
    """Manually approve a KYC request."""
    kyc = get_object_or_404(KYC, pk=pk)
    kyc.is_verified = True
    kyc.verification_score = 100.0  # Force 100% on manual approval
    kyc.save()
    
    # Notify User (Internal notification logic)
    try:
        from notifications.models import send_notification
        send_notification(kyc.user, "KYC Verified!", "Your identity has been manually verified by the portal sentinel.", 'SYSTEM')
    except:
        pass
        
    messages.success(request, f"Identity for {kyc.user.username} has been verified.")
    return redirect('sys_admin:tenant_kyc' if kyc.user.role == 'TENANT' else 'sys_admin:owner_kyc')

@admin_required
def reject_kyc(request, pk):
    """Reject a KYC request."""
    kyc = get_object_or_404(KYC, pk=pk)
    kyc.is_verified = False
    kyc.verification_score = 0.0
    kyc.save()
    
    # Notify User
    try:
        from notifications.models import send_notification
        send_notification(kyc.user, "KYC Rejected", "Your identity verification failed. Please re-upload clear documents.", 'SYSTEM')
    except:
        pass

    messages.warning(request, f"Identity for {kyc.user.username} has been rejected.")
    return redirect('sys_admin:tenant_kyc' if kyc.user.role == 'TENANT' else 'sys_admin:owner_kyc')

@admin_required
def export_accounts_excel(request):
    """Exports all account-related data (User, KYC, ProMembership) to an Excel file."""
    workbook = openpyxl.Workbook()
    
    # --- Sheet 1: Users ---
    sheet_users = workbook.active
    sheet_users.title = "Users"
    headers_users = ['ID', 'Username', 'Email', 'Role', 'Date Joined', 'Is Active', 'Phone Number']
    sheet_users.append(headers_users)
    
    for user in User.objects.all():
        sheet_users.append([
            user.id,
            user.username,
            user.email,
            user.role,
            user.date_joined.replace(tzinfo=None) if user.date_joined else '',
            user.is_active,
            user.phone_number or ''
        ])

    # --- Sheet 2: KYC Data ---
    sheet_kyc = workbook.create_sheet(title="KYC Records")
    headers_kyc = ['User', 'Full Name', 'ID Type', 'ID Number', 'Status', 'Submitted At', 'Tenant Type', 'Occupation', 'Property Address']
    sheet_kyc.append(headers_kyc)
    
    for kyc in KYC.objects.all():
        sheet_kyc.append([
            kyc.user.username,
            kyc.full_name,
            kyc.id_type,
            kyc.id_number,
            'Verified' if kyc.is_verified else 'Pending',
            kyc.submitted_at.replace(tzinfo=None) if kyc.submitted_at else '',
            kyc.tenant_type,
            kyc.occupation,
            kyc.property_address
        ])

    # --- Sheet 3: Pro Memberships ---
    sheet_pro = workbook.create_sheet(title="Pro Memberships")
    headers_pro = ['User', 'Tier', 'Status', 'Expires At']
    sheet_pro.append(headers_pro)
    
    for pro in ProMembership.objects.all():
        sheet_pro.append([
            pro.user.username,
            pro.tier,
            'Active' if pro.is_active else 'Inactive',
            pro.expires_at.replace(tzinfo=None) if pro.expires_at else ''
        ])

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="BrosDen_Accounts_{datetime.now().strftime("%y%m%d_%H%M")}.xlsx"'
    
    workbook.save(response)
    return response

@admin_required
def approve_property(request, pk):
    """Approve a property listing."""
    prop = get_object_or_404(Property, pk=pk)
    prop.is_verified = True
    prop.save()
    
    try:
        from notifications.models import send_notification
        send_notification(prop.owner, "Property Verified!", f"Your property '{prop.title}' has been verified and is now live.", 'SYSTEM')
    except:
        pass
        
    messages.success(request, f"Property '{prop.title}' has been approved.")
    return redirect('sys_admin:listing_requests')

@admin_required
def reject_property(request, pk):
    """Reject a property listing."""
    prop = get_object_or_404(Property, pk=pk)
    prop.is_verified = False
    prop.save()
    
    try:
        from notifications.models import send_notification
        send_notification(prop.owner, "Property Rejected", f"Your property '{prop.title}' verification failed. Please check your listing details.", 'SYSTEM')
    except:
        pass
        
    messages.warning(request, f"Property '{prop.title}' has been rejected.")
    return redirect('sys_admin:listing_requests')
